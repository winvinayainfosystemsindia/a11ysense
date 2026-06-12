import io
import os
import json
import zipfile
import logging
from typing import Optional, List, Dict, Any

from common.config import get_audit_storage_path
from common.utils.event_bus import read_events

logger = logging.getLogger(__name__)

class ReportRepository:
    def get_screenshot_file_path(self, task_id: str, filename: str) -> Optional[str]:
        """Serves a captured defect screenshot path for the given task."""
        reports_dir = get_audit_storage_path(task_id)
        file_path = os.path.join(reports_dir, filename)
        if os.path.exists(file_path):
            return file_path
        return None

    def load_report_json(self, task_id: str) -> Optional[List[Dict[str, Any]]]:
        """Loads compiled report JSON list if it exists."""
        reports_dir = get_audit_storage_path(task_id)
        json_path = os.path.join(reports_dir, f"testcase_report_{task_id}.json")
        if os.path.exists(json_path):
            with open(json_path, "r", encoding="utf-8") as f:
                return json.load(f)
        return None

    def generate_excel_file(self, testcases: List[Dict[str, Any]]) -> io.BytesIO:
        """Generates a styled Excel sheet with Test Cases and Defects tabs."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        wb.remove(wb.active) # Remove default sheet
        
        font_family = "Segoe UI"
        header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        data_font = Font(name=font_family, size=10)
        code_font = Font(name="Consolas", size=9)
        
        header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        zebra_fill = PatternFill(start_color="F2F6FB", end_color="F2F6FB", fill_type="solid")
        pass_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        fail_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        
        border_thin = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        headers = [
            "S.No", "Test Case ID", "Test Case Name", "Page Title", "Page URL", 
            "Rule ID", "Criteria", "Level", "Severity", "Status", "Description", 
            "Expected Result", "Actual Result", "Steps to Reproduce", "Remediation", 
            "Refined By", "Screenshot"
        ]
        
        def populate_sheet(ws, cases):
            ws.append(headers)
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border_thin
                
            ws.row_dimensions[1].height = 28
            
            for row_idx, tc in enumerate(cases, 2):
                steps = tc.get("steps_to_reproduce", "N/A")
                if isinstance(steps, list):
                    steps = "\n".join(str(s) for s in steps)
                else:
                    steps = str(steps) if steps is not None else "N/A"
                    
                remediation = tc.get("remediation", "N/A")
                if isinstance(remediation, list):
                    remediation = "\n".join(str(r) for r in remediation)
                else:
                    remediation = str(remediation) if remediation is not None else "N/A"

                row_data = [
                    row_idx - 1,
                    tc.get("testcase_id", "N/A"),
                    tc.get("testcase_name", "N/A"),
                    tc.get("page_title", "N/A"),
                    tc.get("page_url", "N/A"),
                    tc.get("rule_id", "N/A"),
                    tc.get("criteria", "N/A"),
                    tc.get("level", "N/A"),
                    tc.get("severity", "N/A"),
                    tc.get("status", "N/A"),
                    tc.get("description", "N/A"),
                    tc.get("expected_result", "N/A"),
                    tc.get("actual_result", "N/A"),
                    steps,
                    remediation,
                    tc.get("refined_by", "N/A"),
                    tc.get("screenshot", "N/A")
                ]
                
                ws.append(row_data)
                ws.row_dimensions[row_idx].height = 20
                
                for col_idx, val in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = border_thin
                    
                    if col_idx in [1, 2, 7, 8, 9, 10, 17]:
                        cell.alignment = align_center
                    else:
                        cell.alignment = align_left
                        
                    if col_idx in [2, 6]:
                        cell.font = code_font
                    else:
                        cell.font = data_font
                        
                    if row_idx % 2 == 1:
                        cell.fill = zebra_fill
                        
                    if col_idx == 10:
                        if val == "PASS":
                            cell.fill = pass_fill
                            cell.font = Font(name=font_family, size=10, bold=True, color="155724")
                        elif val == "FAIL":
                            cell.fill = fail_fill
                            cell.font = Font(name=font_family, size=10, bold=True, color="721C24")
                            
                    if col_idx == 9:
                        sev = str(val).lower()
                        if sev in ["critical", "blocker"]:
                            cell.font = Font(name=font_family, size=10, bold=True, color="721C24")
                        elif sev in ["serious", "high"]:
                            cell.font = Font(name=font_family, size=10, bold=True, color="856404")
                            
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(cases) + 1}"
            
            for col in ws.columns:
                max_len = 0
                for cell in col:
                    val_str = str(cell.value or '')
                    lines = val_str.split('\n')
                    longest_line = max(len(l) for l in lines) if lines else 0
                    if longest_line > max_len:
                        max_len = longest_line
                col_letter = get_column_letter(col[0].column)
                adjusted_width = min(max(max_len + 4, 12), 40)
                ws.column_dimensions[col_letter].width = adjusted_width

        ws_tc = wb.create_sheet(title="Test Cases")
        populate_sheet(ws_tc, testcases)
        
        ws_def = wb.create_sheet(title="Defects")
        defects = [tc for tc in testcases if tc.get("status") == "FAIL"]
        populate_sheet(ws_def, defects)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def create_zip_archive(self, task_id: str, testcases: List[Dict[str, Any]], excel_file: io.BytesIO) -> io.BytesIO:
        """Compiles Excel, JSON, and screenshots into a single ZIP archive buffer."""
        reports_dir = get_audit_storage_path(task_id)
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            # 1. Add json file (formatted for readability)
            json_content = json.dumps(testcases, indent=2, ensure_ascii=False)
            zip_file.writestr("report.json", json_content.encode("utf-8"))
            
            # 2. Add excel file
            zip_file.writestr("report.xlsx", excel_file.getvalue())
            
            # 3. Add screenshots folder
            screenshots_added = set()
            for tc in testcases:
                screenshot_filename = tc.get("screenshot")
                if screenshot_filename and screenshot_filename != "N/A":
                    if screenshot_filename not in screenshots_added:
                        src_screenshot_path = os.path.join(reports_dir, screenshot_filename)
                        if os.path.exists(src_screenshot_path):
                            zip_file.write(src_screenshot_path, arcname=f"screenshots/{screenshot_filename}")
                            screenshots_added.add(screenshot_filename)
        zip_buffer.seek(0)
        return zip_buffer

    def read_analyzed_events(self, last_id: str, block_ms: int = 2000) -> list:
        """Polls the 'audit:analyzed' Redis stream for new analysis events."""
        try:
            return read_events("audit:analyzed", last_id=last_id, block_ms=block_ms) or []
        except Exception as e:
            logger.error(f"[ReportRepository] Error reading events: {e}")
            return []

report_repo = ReportRepository()
