import asyncio
import io
from typing import Optional
import sys
import os
import threading
import logging
import time

# Windows proactor loop policy
if sys.platform == 'win32':
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

from common.config import setup_environment, get_storage_path, get_cors_origins, get_audit_storage_path
setup_environment()

from fastapi import FastAPI, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.exceptions import RequestValidationError
from fastapi.responses import JSONResponse, HTMLResponse, FileResponse
from app.schemas.audit import AuditResult
from app.services.report_service import report_service

# Event bus reader utility
from common.utils.event_bus import read_events

logger = logging.getLogger(__name__)

app = FastAPI(title="A11ySense AI Reporting Service")

from common.exceptions.handler import setup_global_exception_handler
setup_global_exception_handler(app, "reporting-service")

from prometheus_fastapi_instrumentator import Instrumentator
Instrumentator().instrument(app).expose(app)


app.add_middleware(
    CORSMiddleware,
    allow_origins=get_cors_origins(),
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError):
    print(f"Validation Error: {exc.errors()}")
    return JSONResponse(
        status_code=422,
        content={"detail": exc.errors()},
    )


@app.post("/generate")
async def generate_report(result: AuditResult, task_id: Optional[str] = None):
    """
    Industry-standard Allure report generation.
    """
    return await report_service.create_audit_report(result, task_id)


@app.get("/report/{task_id}/screenshot/{filename}")
async def get_report_screenshot(task_id: str, filename: str):
    """
    Serves a captured defect screenshot for the given task.
    """
    reports_dir = get_audit_storage_path(task_id)
    file_path = os.path.join(reports_dir, filename)
    if os.path.exists(file_path):
        return FileResponse(file_path, media_type="image/png")
    return JSONResponse(status_code=404, content={"detail": "Screenshot not found"})


@app.get("/report/{task_id}", response_class=HTMLResponse)
async def get_html_report(task_id: str):
    """
    Directly serves a gorgeous, premium HTML render of the compiled audit reports.
    """
    reports_dir = get_audit_storage_path(task_id)
    json_path = os.path.join(reports_dir, f"testcase_report_{task_id}.json")
    
    if os.path.exists(json_path):
        import json
        import html
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                testcases = json.load(f)
        except Exception as e:
            return HTMLResponse(content=f"<h1>Error reading report</h1><p>{str(e)}</p>", status_code=500)

        fail_cases = [tc for tc in testcases if tc.get("status") == "FAIL"]
        pass_cases = [tc for tc in testcases if tc.get("status") == "PASS"]

        html_body = f"<h1>Accessibility Compliance Report</h1>"
        html_body += f"<p><strong>Scan Task ID:</strong> {task_id}</p>"
        html_body += f"<p><strong>Summary:</strong> {len(fail_cases)} Defects Found | {len(pass_cases)} Passed Checks</p>"
        html_body += "<hr/>"

        if fail_cases:
            html_body += f"<h2>Accessibility Defects ({len(fail_cases)})</h2>"
            for idx, tc in enumerate(fail_cases, start=1):
                help_url = tc.get("help_url", "")
                html_snippet = tc.get("html_snippet", "")
                
                testcase_name = str(tc.get('testcase_name', 'N/A'))
                testcase_id = str(tc.get('testcase_id', 'N/A'))
                rule_id = str(tc.get('rule_id', 'N/A'))
                page_title = str(tc.get('page_title', 'N/A'))
                page_url = str(tc.get('page_url', 'N/A'))
                criteria = str(tc.get('criteria', 'N/A'))
                level = str(tc.get('level', 'N/A'))
                severity = str(tc.get('severity', 'N/A'))
                description = str(tc.get('description', 'N/A'))
                business_impact = str(tc.get('business_impact', 'N/A'))
                refined_by = str(tc.get('refined_by', 'N/A'))

                expected_result = tc.get('expected_result', 'N/A')
                expected_result = str(expected_result) if expected_result is not None else 'N/A'
                
                actual_result = tc.get('actual_result', 'N/A')
                actual_result = str(actual_result) if actual_result is not None else 'N/A'

                steps = tc.get('steps_to_reproduce', 'N/A')
                if isinstance(steps, list):
                    steps = "\n".join(str(s) for s in steps)
                else:
                    steps = str(steps) if steps is not None else 'N/A'

                remediation = tc.get('remediation', 'N/A')
                if isinstance(remediation, list):
                    remediation = "\n".join(str(r) for r in remediation)
                else:
                    remediation = str(remediation) if remediation is not None else 'N/A'

                html_body += f"<div style='margin-bottom: 30px; padding: 20px; border: 1px solid #e2e8f0; border-radius: 8px; background-color: #fff;'>"
                html_body += f"<h3>Defect {idx}: {html.escape(testcase_name)}</h3>"
                html_body += "<ul style='padding-left: 20px;'>"
                html_body += f"<li><strong>Test Case ID:</strong> <code>{html.escape(testcase_id)}</code></li>"
                html_body += f"<li><strong>Rule ID:</strong> <code>{html.escape(rule_id)}</code></li>"
                html_body += f"<li><strong>Page Title:</strong> {html.escape(page_title)}</li>"
                html_body += f"<li><strong>URL:</strong> <a href='{html.escape(page_url)}' target='_blank'>{html.escape(page_url)}</a></li>"
                html_body += f"<li><strong>Criteria:</strong> {html.escape(criteria)}</li>"
                html_body += f"<li><strong>Level:</strong> {html.escape(level)}</li>"
                html_body += f"<li><strong>Severity:</strong> <span style='font-weight: 600; color: #991b1b;'>{html.escape(severity)}</span></li>"
                html_body += f"<li><strong>Description:</strong> {html.escape(description)}</li>"
                html_body += f"<li><strong>Business Impact:</strong> {html.escape(business_impact)}</li>"
                html_body += f"<li><strong>Expected Result:</strong> {html.escape(expected_result)}</li>"
                html_body += f"<li><strong>Actual Result:</strong> {html.escape(actual_result)}</li>"
                html_body += f"<li><strong>Steps To Reproduce:</strong> {html.escape(steps)}</li>"
                html_body += f"<li><strong>Remediation:</strong> {html.escape(remediation)}</li>"
                html_body += f"<li><strong>Refined By:</strong> {html.escape(refined_by)}</li>"
                
                # Show token usage sizes
                input_tokens = tc.get("input_tokens", 0)
                output_tokens = tc.get("output_tokens", 0)
                if input_tokens > 0 or output_tokens > 0:
                    html_body += f"<li><strong>AI Token Size:</strong> Input: {input_tokens} | Output: {output_tokens}</li>"

                if help_url:
                    html_body += f"<li><strong>Help URL:</strong> <a href='{html.escape(help_url)}' target='_blank'>{html.escape(help_url)}</a></li>"
                html_body += "</ul>"
                
                if html_snippet:
                    html_body += f"<p><strong>HTML Snippet:</strong></p>"
                    html_body += f"<pre><code class='language-html'>{html.escape(html_snippet)}</code></pre>"
                
                screenshot = tc.get("screenshot")
                if screenshot and screenshot != "N/A":
                    html_body += f"<p><strong>Visual Evidence (Defect Screenshot):</strong></p>"
                    html_body += f"<div style='margin-top: 10px; margin-bottom: 20px;'><img src='/report/{task_id}/screenshot/{html.escape(screenshot)}' alt='Defect Screenshot' style='max-width: 100%; border: 1px solid #cbd5e1; border-radius: 6px; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1);'/></div>"
                
                html_body += "</div>"

        if pass_cases:
            html_body += f"<h2>Passed Compliance Checks ({len(pass_cases)})</h2>"
            html_body += "<table>"
            html_body += "<thead><tr><th>Test Case ID</th><th>Rule ID</th><th>Page Title</th><th>Status</th></tr></thead>"
            html_body += "<tbody>"
            for tc in pass_cases[:50]:  # limit to first 50 passed checks
                html_body += "<tr>"
                html_body += f"<td><code>{html.escape(tc.get('testcase_id', 'N/A'))}</code></td>"
                html_body += f"<td>{html.escape(tc.get('rule_id', 'N/A'))}</td>"
                html_body += f"<td>{html.escape(tc.get('page_title', 'N/A'))}</td>"
                html_body += f"<td style='background-color: #f0fdf4; color: #166534; font-weight: 600;'>PASS</td>"
                html_body += "</tr>"
            html_body += "</tbody></table>"
            if len(pass_cases) > 50:
                html_body += f"<p style='font-style: italic; color: #6b7280; margin-top: 8px;'>Showing first 50 passed checks. Refer to the JSON report for full details.</p>"

        html_page = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Accessibility Audit Report - {task_id}</title>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;700&display=swap" rel="stylesheet">
            <style>
                body {{
                    font-family: 'Outfit', sans-serif;
                    line-height: 1.6;
                    color: #1f2937;
                    max-width: 1000px;
                    margin: 40px auto;
                    padding: 0 20px;
                    background-color: #f3f4f6;
                }}
                li {{
                    margin-bottom: 8px;
                }}
                strong {{
                    color: #111827;
                }}
                .card {{
                    background: white;
                    padding: 30px;
                    border-radius: 12px;
                    box-shadow: 0 10px 15px -3px rgba(0,0, 0, 0.05);
                }}
                h1 {{
                    color: #1e3a8a;
                    border-bottom: 2px solid #e2e8f0;
                    padding-bottom: 10px;
                }}
                h2 {{
                    color: #2563eb;
                    margin-top: 35px;
                    border-bottom: 1px solid #e2e8f0;
                    padding-bottom: 8px;
                }}
                h3 {{
                    color: #1e293b;
                    margin-top: 0;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin: 20px 0;
                    background: white;
                    border-radius: 8px;
                    box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1);
                    overflow: hidden;
                }}
                th {{
                    background-color: #3b82f6;
                    color: white;
                    text-align: left;
                    font-weight: 600;
                    padding: 12px 15px;
                }}
                td {{
                    padding: 12px 15px;
                    border-bottom: 1px solid #e2e8f0;
                    font-size: 0.92rem;
                }}
                tr:hover td {{
                    background-color: rgba(0, 0, 0, 0.01);
                }}
                pre {{
                    background-color: #f8fafc;
                    border: 1px solid #e2e8f0;
                    padding: 15px;
                    border-radius: 6px;
                    overflow-x: auto;
                    font-family: monospace;
                    font-size: 0.85rem;
                }}
            </style>
        </head>
        <body>
            <div class="card">
                {html_body}
            </div>
        </body>
        </html>
        """
        return HTMLResponse(content=html_page)
        
    return HTMLResponse(content="<h1>Report Not Found</h1><p>The report for this task ID does not exist yet.</p>", status_code=404)


@app.get("/report/{task_id}/export")
async def export_report(task_id: str):
    """
    Generates and exports a ZIP archive containing:
      - report.json: Full audit report JSON
      - report.xlsx: Styled Excel sheet with Test Cases & Defects tabs
      - screenshots/: Folder of referenced screenshot files
    """
    import io
    import json
    import zipfile
    from fastapi.responses import StreamingResponse
    
    reports_dir = get_audit_storage_path(task_id)
    json_path = os.path.join(reports_dir, f"testcase_report_{task_id}.json")
    
    if not os.path.exists(json_path):
        return JSONResponse(status_code=404, content={"detail": "Report JSON not found"})
        
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            testcases = json.load(f)
    except Exception as e:
        return JSONResponse(status_code=500, content={"detail": f"Error reading report: {str(e)}"})
        
    # Generate Excel report
    try:
        excel_file = generate_excel_report(testcases)
    except Exception as e:
        logger.exception("Failed to generate Excel report")
        return JSONResponse(status_code=500, content={"detail": f"Error generating Excel report: {str(e)}"})
        
    # Generate Zip archive in memory
    zip_buffer = io.BytesIO()
    try:
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            # 1. Add json file (formatted for readability)
            json_content = json.dumps(testcases, indent=2, ensure_ascii=False)
            zip_file.writestr("report.json", json_content.encode("utf-8"))
            
            # 2. Add excel file
            zip_file.writestr("report.xlsx", excel_file.getvalue())
            
            # 3. Add screenshots folder
            screenshots_added = set()
            for tc in testcases:
                screenshot_filename = tc.get("screenshot")
                if screenshot_filename and screenshot_filename != "N/A":
                    if screenshot_filename not in screenshots_added:
                        src_screenshot_path = os.path.join(reports_dir, screenshot_filename)
                        if os.path.exists(src_screenshot_path):
                            zip_file.write(src_screenshot_path, arcname=f"screenshots/{screenshot_filename}")
                            screenshots_added.add(screenshot_filename)
    except Exception as e:
        logger.exception("Failed to compile ZIP archive")
        return JSONResponse(status_code=500, content={"detail": f"Error compiling ZIP: {str(e)}"})
        
    zip_buffer.seek(0)
    
    headers = {
        'Content-Disposition': f'attachment; filename="audit_report_{task_id}.zip"'
    }
    return StreamingResponse(zip_buffer, media_type="application/zip", headers=headers)


def generate_excel_report(testcases) -> io.BytesIO:
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet
    
    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10)
    code_font = Font(name="Consolas", size=9)
    
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    zebra_fill = PatternFill(start_color="F2F6FB", end_color="F2F6FB", fill_type="solid")
    pass_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    fail_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    
    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    headers = [
        "S.No", "Test Case ID", "Test Case Name", "Page Title", "Page URL", 
        "Rule ID", "Criteria", "Level", "Severity", "Status", "Description", 
        "Expected Result", "Actual Result", "Steps to Reproduce", "Remediation", 
        "Refined By", "Screenshot"
    ]
    
    def populate_sheet(ws, cases):
        ws.append(headers)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border_thin
            
        ws.row_dimensions[1].height = 28
        
        for row_idx, tc in enumerate(cases, 2):
            steps = tc.get("steps_to_reproduce", "N/A")
            if isinstance(steps, list):
                steps = "\n".join(str(s) for s in steps)
            else:
                steps = str(steps) if steps is not None else "N/A"
                
            remediation = tc.get("remediation", "N/A")
            if isinstance(remediation, list):
                remediation = "\n".join(str(r) for r in remediation)
            else:
                remediation = str(remediation) if remediation is not None else "N/A"

            row_data = [
                row_idx - 1,
                tc.get("testcase_id", "N/A"),
                tc.get("testcase_name", "N/A"),
                tc.get("page_title", "N/A"),
                tc.get("page_url", "N/A"),
                tc.get("rule_id", "N/A"),
                tc.get("criteria", "N/A"),
                tc.get("level", "N/A"),
                tc.get("severity", "N/A"),
                tc.get("status", "N/A"),
                tc.get("description", "N/A"),
                tc.get("expected_result", "N/A"),
                tc.get("actual_result", "N/A"),
                steps,
                remediation,
                tc.get("refined_by", "N/A"),
                tc.get("screenshot", "N/A")
            ]
            
            ws.append(row_data)
            ws.row_dimensions[row_idx].height = 20
            
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border_thin
                
                if col_idx in [1, 2, 7, 8, 9, 10, 17]:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left
                    
                if col_idx in [2, 6]:
                    cell.font = code_font
                else:
                    cell.font = data_font
                    
                if row_idx % 2 == 1:
                    cell.fill = zebra_fill
                    
                if col_idx == 10:
                    if val == "PASS":
                        cell.fill = pass_fill
                        cell.font = Font(name=font_family, size=10, bold=True, color="155724")
                    elif val == "FAIL":
                        cell.fill = fail_fill
                        cell.font = Font(name=font_family, size=10, bold=True, color="721C24")
                        
                if col_idx == 9:
                    sev = str(val).lower()
                    if sev in ["critical", "blocker"]:
                        cell.font = Font(name=font_family, size=10, bold=True, color="721C24")
                    elif sev in ["serious", "high"]:
                        cell.font = Font(name=font_family, size=10, bold=True, color="856404")
                        
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(cases) + 1}"
        
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                lines = val_str.split('\n')
                longest_line = max(len(l) for l in lines) if lines else 0
                if longest_line > max_len:
                    max_len = longest_line
            col_letter = get_column_letter(col[0].column)
            adjusted_width = min(max(max_len + 4, 12), 40)
            ws.column_dimensions[col_letter].width = adjusted_width

    ws_tc = wb.create_sheet(title="Test Cases")
    populate_sheet(ws_tc, testcases)
    
    ws_def = wb.create_sheet(title="Defects")
    defects = [tc for tc in testcases if tc.get("status") == "FAIL"]
    populate_sheet(ws_def, defects)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── Reporting Stream Worker Thread ──────────────────────────────────────────

def start_reporting_worker():
    """
    Spawns the background ReportingWorker thread.
    """
    thread = threading.Thread(target=_run_reporting_worker, daemon=True)
    thread.start()
    logger.info("Reporting Worker: Background thread spawned successfully.")


def _run_reporting_worker():
    """
    Blocking thread loop that polls the 'audit:analyzed' Redis stream.
    """
    time.sleep(2.0)
    logger.info("Reporting Worker: Loop started. Listening to stream 'audit:analyzed'...")
    
    # Listen to new events created after startup
    last_id = "$"
    
    while True:
        try:
            events = read_events("audit:analyzed", last_id=last_id, block_ms=2000)
            if not events:
                time.sleep(1.0)
                continue
            for msg_id, payload in events:
                last_id = msg_id
                
                task_id = payload.get("task_id")
                result_dict = payload.get("result")
                if not task_id or not result_dict:
                    continue
                
                logger.info(f"Reporting Worker: Received finished audit data for task {task_id}")
                
                # Compile report asynchronously in a dedicated loop
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                try:
                    result = AuditResult(**result_dict)
                    loop.run_until_complete(report_service.create_audit_report(result, task_id))
                    logger.info(f"Reporting Worker: Successfully generated Allure report for task {task_id}")
                except Exception as report_err:
                    logger.error(f"Reporting Worker: Report generation failed for task {task_id}: {str(report_err)}")
                finally:
                    loop.close()
                    
        except Exception as e:
            logger.error(f"Reporting Worker: Loop error: {str(e)}")
            time.sleep(5.0)


@app.on_event("startup")
async def startup_event():
    """Start background Reporting Stream listener thread on service launch."""
    start_reporting_worker()
